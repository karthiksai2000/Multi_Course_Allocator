"""
Streamlit Web Application for Dual Course Allocation System
With Manual Section Definition
"""
import streamlit as st
import pandas as pd
import os
import re
from pathlib import Path
from datetime import datetime
import sys
from io import BytesIO

sys.path.insert(0, str(Path(__file__).parent.parent))

from app.models import AllocationConfig, Section
from app.excel_handler import ExcelHandler
from app.allocator import AllocationEngine

# Set page config
st.set_page_config(
    page_title="Dual Course Allocation System",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Custom CSS
st.markdown("""
    <style>
    .main {
        padding: 2rem;
    }
    .step-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 10px;
        margin-bottom: 1.5rem;
    }
    .allocated-card { background: #d4edda; padding: 1rem; border-radius: 5px; border-left: 5px solid #28a745; }
    .unallocated-card { background: #f8d7da; padding: 1rem; border-radius: 5px; border-left: 5px solid #dc3545; }
    .stats-card { background: #d1ecf1; padding: 1rem; border-radius: 5px; border-left: 5px solid #17a2b8; }
    </style>
""", unsafe_allow_html=True)

# Initialize session state
if 'step' not in st.session_state:
    st.session_state.step = 1
if 'config' not in st.session_state:
    st.session_state.config = None
if 'students' not in st.session_state:
    st.session_state.students = []
if 'results' not in st.session_state:
    st.session_state.results = None
if 'stats' not in st.session_state:
    st.session_state.stats = None
if 'paired_sections' not in st.session_state:
    st.session_state.paired_sections = []
if 'section_mode' not in st.session_state:
    st.session_state.section_mode = None  # None, "manual", or "automatic"
if 'manual_sections' not in st.session_state:
    st.session_state.manual_sections = []
if 'auto_planned_sections' not in st.session_state:
    st.session_state.auto_planned_sections = []
if 'auto_num_sections' not in st.session_state:
    st.session_state.auto_num_sections = None
if 'allocation_criteria' not in st.session_state:
    st.session_state.allocation_criteria = None  # "CGPA" or "Timestamp"
if 'auto_sections_edited' not in st.session_state:
    st.session_state.auto_sections_edited = False  # Track if user edited the auto plan

# Page title
st.title("📚 Dual Course Allocation System")
st.markdown("Fair allocation with manual section definition")

# Progress indicator
st.markdown("---")
progress_cols = st.columns(4)
steps = ["1️⃣ Upload", "2️⃣ Sections", "3️⃣ Allocate", "4️⃣ Results"]

for i, step_name in enumerate(steps):
    with progress_cols[i]:
        if i < st.session_state.step - 1:
            st.markdown(f"✓ {step_name}", unsafe_allow_html=True)
        elif i == st.session_state.step - 1:
            st.markdown(f"**✅ {step_name}**", unsafe_allow_html=True)
        else:
            st.markdown(f"○ {step_name}", unsafe_allow_html=True)

st.markdown("---")

# ====================
# STEP 1: UPLOAD FILES
# ====================
if st.session_state.step == 1:
    st.markdown("""
    <div class="step-header">
        <h2>Step 1️⃣: Upload Files</h2>
        <p>Upload course configuration and student data</p>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📖 Course Configuration")
        st.info("Upload Excel file with courses in columns: Course Name, Group, Prerequisites")
        config_file = st.file_uploader("Upload courses (Excel)", type=["xlsx"], key="config_upload")
        
        if config_file:
            try:
                st.session_state.config = ExcelHandler.read_courses_from_excel(config_file)
                g1 = len([c for c in st.session_state.config.courses.values() if c.group.value == "G1"])
                g2 = len([c for c in st.session_state.config.courses.values() if c.group.value == "G2"])
                st.success(f"✅ Loaded {g1} G1 courses + {g2} G2 courses")
                
                # Diagnostic: Show all loaded courses with their groups
                with st.expander("🔍 View Course Details"):
                    g1_courses = [c for c in st.session_state.config.courses.values() if c.group.value == "G1"]
                    g2_courses = [c for c in st.session_state.config.courses.values() if c.group.value == "G2"]
                    
                    st.write("**Group 1 Courses:**")
                    for c in sorted(g1_courses, key=lambda x: x.course_name):
                        st.caption(f"✅ {c.course_name}")
                    
                    st.write("**Group 2 Courses:**")
                    for c in sorted(g2_courses, key=lambda x: x.course_name):
                        st.caption(f"✅ {c.course_name}")
                        
            except Exception as e:
                st.error(f"Error: {str(e)}")
    
    with col2:
        st.subheader("👥 Student Data")
        st.info("Upload Excel file with columns: Name, Registration, CGPA, G1 Preferences, G2 Preferences")
        student_file = st.file_uploader("Upload students (Excel)", type=["xlsx"], key="student_upload")
        
        if student_file:
            try:
                st.session_state.students, skip_errors = ExcelHandler.read_students_from_excel(student_file)
                loaded_count = len(st.session_state.students)
                st.success(f"✅ Loaded {loaded_count} students")
                
                # Show warning if any students were skipped
                if skip_errors:
                    skipped_count = len(skip_errors)
                    st.warning(f"""
                    ⚠️ **{skipped_count} student(s) were skipped due to data issues:**
                    """)
                    
                    # Show first 5 errors
                    for error in skip_errors[:5]:
                        st.caption(f"❌ {error}")
                    
                    if len(skip_errors) > 5:
                        st.caption(f"... and {len(skip_errors) - 5} more")
                    
                    with st.expander("📋 View All Errors"):
                        for error in skip_errors:
                            st.text(error)
                            
            except Exception as e:
                st.error(f"Error: {str(e)}")
    
    # Display course combinations analysis
    if st.session_state.config and st.session_state.students:
        st.markdown("---")
        st.subheader("📊 Course Combination Analysis")
        st.info("📌 View which G1+G2 combinations students prefer to better allocate sections")
        
        # Create allocator just for analysis
        try:
            engine = AllocationEngine(st.session_state.config)
            combination_data = engine.analyze_course_combinations(st.session_state.students)
            
            if combination_data["status"] == "success":
                # Summary metrics
                col_total, col_unique, col_top = st.columns(3)
                
                with col_total:
                    st.metric("Total Students", combination_data["total_students"])
                
                with col_unique:
                    st.metric("Unique Combinations", combination_data["unique_combinations"])
                
                with col_top:
                    if combination_data["top_combination"]:
                        top = combination_data["top_combination"]
                        st.metric(
                            "Most Popular",
                            f"{top['g1_course']} + {top['g2_course']}",
                            f"{top['student_count']} students"
                        )
                
                # Display combinations table
                st.markdown("### Demand by Combination (Sorted by Popularity)")
                
                # Create DataFrame for better display
                combinations_df = pd.DataFrame([
                    {
                        "Rank": i + 1,
                        "G1 Course": comb["g1_course"],
                        "G2 Course": comb["g2_course"],
                        "Student Count": comb["student_count"],
                        "Percentage": f"{comb['percentage']}%"
                    }
                    for i, comb in enumerate(combination_data["combinations"])
                ])
                
                st.dataframe(
                    combinations_df,
                    use_container_width=True,
                    hide_index=True
                )
                
                # Recommendations
                st.markdown("### 💡 Recommendations for Section Creation")
                recommendations = []
                for i, comb in enumerate(combination_data["combinations"][:5], 1):
                    if comb["percentage"] >= 5:  # Show if at least 5% of students
                        recommendation = f"**{i}. {comb['g1_course']} + {comb['g2_course']}**: {comb['student_count']} students ({comb['percentage']}%) - Create {'2-3' if comb['student_count'] > 50 else '1-2'} sections"
                        recommendations.append(recommendation)
                
                if recommendations:
                    for rec in recommendations:
                        st.markdown(f"✅ {rec}")
                else:
                    st.info("No dominant combinations found. Consider creating sections based on all combinations.")
                
                # Chart visualization
                st.markdown("### 📈 Student Distribution Across Combinations")
                
                chart_data = pd.DataFrame({
                    "Combination": [f"{c['g1_course']}\n+\n{c['g2_course']}" for c in combination_data["combinations"]],
                    "Students": [c["student_count"] for c in combination_data["combinations"]]
                })
                
                st.bar_chart(
                    chart_data.set_index("Combination"),
                    height=400
                )
        
        except Exception as e:
            st.error(f"Error analyzing combinations: {str(e)}")
    
    # Proceed button
    if st.session_state.config and st.session_state.students:
        st.markdown("---")
        btn_col1, btn_col2 = st.columns([0.5, 0.5])
        
        with btn_col1:
            if st.button("🔄 Clear All Data", use_container_width=True):
                st.session_state.config = None
                st.session_state.students = []
                st.rerun()
        
        with btn_col2:
            if st.button("✅ Proceed to Step 2", use_container_width=True, type="primary"):
                st.session_state.step = 2
                st.rerun()

# ====================
# STEP 2: DEFINE SECTIONS (Manual or Automatic)
# ====================
elif st.session_state.step == 2:
    st.markdown("""
    <div class="step-header">
        <h2>Step 2️⃣: Define Sections</h2>
        <p>Choose: Manual configuration or Automatic generation</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Choice tabs: Manual vs Automatic
    tab1, tab2 = st.tabs(["🤖 Automatic (Recommended)", "✏️ Manual (Expert)"])
    
    # ====================
    # AUTOMATIC MODE
    # ====================
    with tab1:
        st.markdown("### 🤖 Automatic Section Generation")
        st.success("""
        ✨ **System-optimized section creation**
        - You decide: How many total sections?
        - You decide: Allocation criteria (CGPA or Timestamp)
        - System distributes students fairly across them
        """)
        
        # Analyze demand
        combination_data = AllocationEngine(st.session_state.config).analyze_course_combinations(
            st.session_state.students
        )
        
        # Show current demand summary
        st.subheader(f"📊 Current Demand Analysis")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Students", len(st.session_state.students))
        with col2:
            st.metric("Unique Combinations", combination_data["unique_combinations"])
        with col3:
            st.metric("Avg Students/Combo", f"{len(st.session_state.students) // combination_data['unique_combinations']}")
        
        # ==== INPUT 1: Allocation Criteria ====
        st.divider()
        st.subheader("❓ Allocation Priority - How to prioritize students?")
        st.info("Choose the criteria for allocating students to their preferred sections.")
        
        allocation_criteria = st.radio(
            "Priority Criteria",
            options=["CGPA (Merit-based)", "Timestamp (First-come-first-serve)"],
            horizontal=True,
            key="auto_criteria"
        )
        
        if allocation_criteria == "CGPA (Merit-based)":
            st.caption("✅ Students with higher CGPA get priority allocation")
        else:
            st.caption("✅ Students who registered earlier get priority allocation")
        
        # ==== INPUT 2: Number of Sections ====
        st.divider()
        st.subheader("❓ How Many Sections Do You Want?")
        st.info("""
        Choose total number of sections. System will distribute them across combinations based on demand.
        
        **TIP:** For 92 students with 16 combinations:
        - 2-3 sections: Most students share sections (multi-pass fallback handles overflow)
        - 5-8 sections: Dedicated sections for high-demand combos
        - 10+ sections: One section per high-demand combo
        """)
        
        # Better default: sqrt(students) captures ~9-10 for 92 students
        import math
        default_sections = max(2, min(5, math.ceil(math.sqrt(len(st.session_state.students)))))
        
        num_sections_input = st.number_input(
            "Total Sections to Create",
            min_value=1,
            max_value=50,
            value=default_sections,
            key="auto_sections_input"
        )
        
        # Reset edits if user changes number of sections
        if st.session_state.get("auto_num_sections") != num_sections_input:
            st.session_state.auto_sections_edited = False
            st.session_state.auto_planned_sections = None
        
        st.caption(f"📊 Avg capacity per section: {len(st.session_state.students) // num_sections_input} students")
        
        # Once user sets both inputs, calculate and show the plan
        if num_sections_input and allocation_criteria:
            st.divider()
            st.subheader(f"📋 Generated Section Plan ({num_sections_input} sections)")
            criteria_display = allocation_criteria.split(" (")[0]  # Extract just "CGPA" or "Timestamp"
            st.write(f"**Total Sections**: {num_sections_input} | **Priority Criteria**: {criteria_display}")
            
            # Algorithm: Select top N combos by demand and allocate one section each
            planned_sections = []
            
            # Sort combinations by student count (descending)
            sorted_combos = sorted(
                combination_data["combinations"],
                key=lambda x: x["student_count"],
                reverse=True
            )
            
            total_students = len(st.session_state.students)
            
            # Simple capacity calculation: equal division
            base_capacity = max(50, int(total_students / num_sections_input))
            
            sections_to_create = min(num_sections_input, len(sorted_combos))
            
            for idx in range(sections_to_create):
                combo = sorted_combos[idx]
                g1 = combo["g1_course"]
                g2 = combo["g2_course"]
                count = combo["student_count"]
                percentage = combo["percentage"]
                
                planned_sections.append({
                    "Combination": f"{g1} + {g2}",
                    "Demand": f"{count} students ({percentage}%)",
                    "Section": f"Section {idx + 1}",
                    "Capacity": base_capacity
                })
            
            # Display the plan
            if planned_sections:
                df_planned = pd.DataFrame(planned_sections)
                st.dataframe(df_planned, use_container_width=True, hide_index=True)
                
                # Final metrics
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Students", len(st.session_state.students))
                with col2:
                    st.metric("Combinations Available", combination_data["unique_combinations"])
                with col3:
                    st.metric("Sections Created", len(planned_sections))
                with col4:
                    total_cap = sum(s["Capacity"] for s in planned_sections)
                    st.metric("Total Capacity", total_cap)
                
                # Explain allocation strategy
                criteria_display = allocation_criteria.split(" (")[0]  # Extract just "CGPA" or "Timestamp"
                st.info(f"""
                ✨ **Allocation Strategy**:
                - **{len(planned_sections)} sections** created for **{combination_data['unique_combinations']} combinations**
                - High-demand combos get dedicated sections
                - Overflow students use **multi-pass fallback** → next preferences
                - All allocations use **{criteria_display}** priority
                """)
                
                # ==== EDIT SECTIONS ====
                with st.expander("✏️ Edit Sections - Customize the plan"):
                    edit_col1, edit_col2 = st.columns([2, 1])
                    
                    with edit_col1:
                        st.info("Modify course combinations and capacities before allocation")
                    
                    with edit_col2:
                        if st.session_state.get("auto_sections_edited", False):
                            st.success("✏️ Customized", icon="✅")
                    
                    # Get available courses
                    all_g1_courses = sorted([c.course_name for c in st.session_state.config.courses.values() if c.group.value == "G1"])
                    all_g2_courses = sorted([c.course_name for c in st.session_state.config.courses.values() if c.group.value == "G2"])
                    
                    edited_sections = []
                    
                    for idx, section in enumerate(planned_sections):
                        st.markdown(f"**Section {idx + 1}**")
                        
                        col1, col2, col3 = st.columns([2, 2, 1])
                        
                        with col1:
                            # Parse current G1 and G2 from combination string
                            combo_parts = section["Combination"].split(" + ")
                            current_g1 = combo_parts[0] if len(combo_parts) > 0 else all_g1_courses[0]
                            
                            edited_g1 = st.selectbox(
                                "Group 1 Course",
                                options=all_g1_courses,
                                index=all_g1_courses.index(current_g1) if current_g1 in all_g1_courses else 0,
                                key=f"edit_g1_{idx}"
                            )
                        
                        with col2:
                            current_g2 = combo_parts[1] if len(combo_parts) > 1 else all_g2_courses[0]
                            edited_g2 = st.selectbox(
                                "Group 2 Course",
                                options=all_g2_courses,
                                index=all_g2_courses.index(current_g2) if current_g2 in all_g2_courses else 0,
                                key=f"edit_g2_{idx}"
                            )
                        
                        with col3:
                            edited_capacity = st.number_input(
                                "Capacity",
                                min_value=20,
                                max_value=200,
                                value=section["Capacity"],
                                key=f"edit_cap_{idx}"
                            )
                        
                        edited_sections.append({
                            "Combination": f"{edited_g1} + {edited_g2}",
                            "Demand": section["Demand"],
                            "Section": section["Section"],
                            "Capacity": edited_capacity
                        })
                        
                        st.divider()
                    
                    # Apply changes buttons
                    btn_col1, btn_col2 = st.columns([1, 1])
                    
                    with btn_col1:
                        if st.button("✅ Apply Changes", use_container_width=True, key="apply_edits"):
                            st.session_state.auto_planned_sections = edited_sections
                            st.session_state.auto_sections_edited = True
                            st.success("✅ Sections updated! Ready to use.")
                            st.rerun()
                    
                    with btn_col2:
                        if st.button("🔄 Reset to Auto", use_container_width=True, key="reset_edits"):
                            st.session_state.auto_sections_edited = False
                            st.session_state.auto_planned_sections = None
                            st.info("✅ Reset to auto-generated sections")
                            st.rerun()
                
                # Store the planned sections in session state ONLY if not edited
                if not st.session_state.get("auto_sections_edited", False):
                    st.session_state.auto_planned_sections = planned_sections
                st.session_state.auto_num_sections = num_sections_input
        
        # Button to proceed with automatic mode
        if st.button("✅ Use Automatic Mode", use_container_width=True, type="primary", key="auto_mode"):
            # Extract criteria value (remove the description part)
            criteria_value = allocation_criteria.split(" (")[0]  # e.g., "CGPA" or "Timestamp"
            st.session_state.allocation_criteria = criteria_value
            
            # Convert planned sections to manual format for allocation
            # This preserves any edits the user made
            planned = st.session_state.auto_planned_sections if st.session_state.auto_planned_sections else planned_sections
            
            manual_format = []
            for idx, section in enumerate(planned):
                combo_parts = section["Combination"].split(" + ")
                g1 = combo_parts[0].strip() if len(combo_parts) > 0 else ""
                g2 = combo_parts[1].strip() if len(combo_parts) > 1 else ""
                
                manual_format.append({
                    'section_number': idx + 1,
                    'g1_course': g1,
                    'g2_course': g2,
                    'capacity': section["Capacity"]
                })
            
            st.session_state.section_mode = "automatic"
            st.session_state.manual_sections = manual_format
            st.rerun()
    
    
    # ====================
    # MANUAL MODE
    # ====================
    with tab2:
        st.markdown("### ✏️ Manual Section Definition")
        st.warning("""
        ⚙️ **Define sections manually**
        - You control exact sections created
        - You set course combinations
        - You set capacities
        - Useful for special requirements
        """)
        
        # ==== Allocation Criteria for Manual Mode ====
        st.divider()
        st.subheader("❓ Allocation Priority - How to prioritize students?")
        st.info("Choose the criteria for allocating students to their preferred sections.")
        
        manual_allocation_criteria = st.radio(
            "Priority Criteria",
            options=["CGPA (Merit-based)", "Timestamp (First-come-first-serve)"],
            horizontal=True,
            key="manual_criteria"
        )
        
        if manual_allocation_criteria == "CGPA (Merit-based)":
            st.caption("✅ Students with higher CGPA get priority allocation")
        else:
            st.caption("✅ Students who registered earlier get priority allocation")
        
        # ==== Manual Section Definition ====
        st.divider()
        st.subheader("📋 Define Your Sections")
        
        # Get available courses
        g1_courses = [c.course_name for c in st.session_state.config.courses.values() if c.group.value == "G1"]
        g2_courses = [c.course_name for c in st.session_state.config.courses.values() if c.group.value == "G2"]
        
        col_g1, col_g2, col_cap, col_add = st.columns([2, 2, 1, 1])
        
        with col_g1:
            selected_g1 = st.selectbox("📖 Group 1 Course", g1_courses, key="manual_g1")
        
        with col_g2:
            selected_g2 = st.selectbox("📖 Group 2 Course", g2_courses, key="manual_g2")
        
        with col_cap:
            capacity = st.number_input("Capacity", min_value=10, max_value=200, value=70, key="manual_cap")
        
        with col_add:
            st.markdown("")
            if st.button("➕ Add Section", use_container_width=True):
                new_section = {
                    'section_number': len(st.session_state.manual_sections) + 1,
                    'g1_course': selected_g1,
                    'g2_course': selected_g2,
                    'capacity': capacity
                }
                st.session_state.manual_sections.append(new_section)
                st.success(f"✅ Added: {selected_g1} + {selected_g2} (Capacity: {capacity})")
        
        # Display defined sections
        if st.session_state.manual_sections:
            st.subheader("📋 Manually Defined Sections")
            df_manual = pd.DataFrame(st.session_state.manual_sections)
            st.dataframe(df_manual, use_container_width=True)
            
            # Clear button
            if st.button("🗑️ Clear All Sections", use_container_width=True):
                st.session_state.manual_sections = []
                st.rerun()
        else:
            st.info("No sections defined yet. Add sections above.")
        
        # Proceed with manual
        if st.session_state.manual_sections:
            if st.button("✅ Use Manual Mode", use_container_width=True, type="primary", key="manual_mode"):
                # Extract criteria value (remove the description part)
                criteria_value = manual_allocation_criteria.split(" (")[0]  # e.g., "CGPA" or "Timestamp"
                st.session_state.allocation_criteria = criteria_value
                st.session_state.section_mode = "manual"
                st.rerun()
    
    # ====================
    # NAVIGATION (Bottom)
    # ====================
    st.divider()
    nav_col1, nav_col2, nav_col3 = st.columns([1, 2, 1])
    
    with nav_col1:
        if st.button("⬅️ Back", use_container_width=True):
            st.session_state.step = 1
            st.session_state.section_mode = None
            st.rerun()
    
    with nav_col2:
        if st.session_state.section_mode:
            if st.button("✅ Proceed to Allocation", use_container_width=True, type="primary"):
                st.session_state.step = 3
                st.rerun()
        else:
            st.info("👆 Select a mode above to proceed")
    
    with nav_col3:
        st.markdown("")  # Spacing

# ====================
# STEP 3: RUN ALLOCATION
# ====================
elif st.session_state.step == 3:
    st.markdown("""
    <div class="step-header">
        <h2>Step 3️⃣: Run Allocation</h2>
        <p>Execute allocation algorithm</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Show mode and summary
    criteria_display = st.session_state.allocation_criteria if st.session_state.allocation_criteria else "Not Selected"
    
    if st.session_state.section_mode == "automatic":
        if st.session_state.auto_sections_edited:
            st.info(f"""
            ✨ **Automatic Mode with Custom Edits**
            1. System generated {st.session_state.auto_num_sections} sections based on demand
            2. You customized the course combinations and capacities
            3. Allocates fairly using: **{criteria_display}**
            4. Multi-pass fallback to alternatives
            """)
        else:
            st.info(f"""
            ✨ **Automatic Mode**
            1. System analyzes student demand
            2. Creates {st.session_state.auto_num_sections} sections based on distribution
            3. Allocates fairly using: **{criteria_display}**
            4. Multi-pass fallback to alternatives
            """)
    else:
        st.info(f"""
        ⚙️ **Manual Mode**
        1. Using your {len(st.session_state.manual_sections)} defined sections
        2. Allocates fairly using: **{criteria_display}**
        3. Multi-pass fallback to alternatives
        """)
    
    # Navigation and execution
    nav_col1, nav_col2 = st.columns(2)
    
    with nav_col1:
        if st.button("⬅️ Back to Sections", use_container_width=True):
            st.session_state.step = 2
            st.session_state.auto_sections_edited = False  # Reset edit flag when going back
            st.rerun()
    
    with nav_col2:
        mode_label = "Automatic" if st.session_state.section_mode == "automatic" else "Manual"
        if st.button(f"🚀 Execute ({mode_label})", use_container_width=True, type="primary"):
            with st.spinner("🔄 Running allocation..."):
                try:
                    engine = AllocationEngine(st.session_state.config)
                    
                    if st.session_state.section_mode == "automatic":
                        # AUTOMATIC MODE: Check if user customized the sections
                        if st.session_state.get("auto_sections_edited", False):
                            # User customized sections - use manual allocation with custom sections
                            results, stats = engine.allocate_with_manual_sections(
                                st.session_state.students,
                                st.session_state.manual_sections,
                                criterion=st.session_state.allocation_criteria
                            )
                        else:
                            # Use auto-generated sections
                            results, stats = engine.allocate(
                                st.session_state.students,
                                criterion=st.session_state.allocation_criteria,
                                target_sections=st.session_state.auto_num_sections
                            )
                        st.session_state.paired_sections = engine.paired_sections
                    else:
                        # MANUAL MODE: Use user-defined sections
                        results, stats = engine.allocate_with_manual_sections(
                            st.session_state.students,
                            st.session_state.manual_sections,
                            criterion=st.session_state.allocation_criteria
                        )
                        st.session_state.paired_sections = engine.paired_sections
                    
                    st.session_state.results = results
                    st.session_state.stats = stats
                    st.session_state.step = 4
                    st.success("✅ Allocation completed!")
                    st.rerun()
                    
                except Exception as e:
                    st.error(f"❌ Error during allocation: {str(e)}")
                    st.error(str(e))

# ====================
# STEP 4: RESULTS
# ====================
elif st.session_state.step == 4:
    st.markdown("""
    <div class="step-header">
        <h2>Step 4️⃣: Allocation Results</h2>
        <p>View sections and allocation results</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Show sections (generated or manual)
    if st.session_state.section_mode == "automatic":
        if st.session_state.get("auto_sections_edited", False):
            mode_label = "✏️ Custom (Auto-Based)"
            info_text = "✏️ These sections are auto-generated but were customized before allocation"
        else:
            mode_label = "🤖 Auto-Generated"
            info_text = "✨ These sections were created automatically based on demand analysis"
    else:
        mode_label = "✏️ Manual"
        info_text = "⚙️ These are the sections you defined for allocation"
    
    st.subheader(f"🎯 {mode_label} Sections")
    st.info(info_text)
    
    if hasattr(st.session_state, 'paired_sections') and st.session_state.paired_sections:
        df_sections_created = pd.DataFrame([{
            'Section': s.get_label(),
            'G1 Course': s.g1_course,
            'G2 Course': s.g2_course,
            'Capacity': s.capacity,
            'Enrolled': s.enrolled,
            'Available': s.available_seats(),
            'Load %': f"{s.get_load_percentage():.1f}%"
        } for s in st.session_state.paired_sections])
        
        st.dataframe(df_sections_created, use_container_width=True)
    
    st.divider()
    
    # Statistics
    st.subheader("📈 Statistics")
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Total Students", st.session_state.stats.total_students)
    
    with col2:
        st.metric("✅ Allocated", st.session_state.stats.allocated_students, 
                 delta=f"{st.session_state.stats.allocation_rate*100:.1f}%")
    
    with col3:
        st.metric("❌ Unallocated", st.session_state.stats.unallocated_students)
    
    with col4:
        st.metric("📊 Success Rate", f"{st.session_state.stats.allocation_rate*100:.1f}%")

    allocated = [r for r in st.session_state.results if r.allocated]
    unallocated = [r for r in st.session_state.results if not r.allocated]

    st.divider()
    st.subheader("🤖 AI Assistant")
    st.caption("Ask questions on current dual-course results. The assistant is visible directly in Step 4.")

    if 'dual_ai_question' not in st.session_state:
        st.session_state.dual_ai_question = "Show allocated students in Section 1"

    dual_examples = [
        "Show allocated students in Section 1",
        "List unallocated students",
        "Give me allocation summary",
        "Show all students allocated to AI",
        "Show all students allocated to Cybersecurity",
    ]

    example_dual = st.selectbox("Example prompts", dual_examples, key='dual_ai_examples')

    q_col1, q_col2 = st.columns([3, 1])
    with q_col1:
        st.session_state.dual_ai_question = st.text_input(
            "Ask AI",
            value=st.session_state.dual_ai_question,
            key='dual_ai_input'
        )
    with q_col2:
        st.write("")
        if st.button("Use Example", key='dual_ai_use_example', use_container_width=True):
            st.session_state.dual_ai_question = example_dual
            st.rerun()

    if st.button("Ask AI", key='dual_ai_ask'):
        question = str(st.session_state.dual_ai_question or "").strip()
        lower = question.lower()

        if not question:
            st.session_state.dual_ai_answer = {
                'kind': 'text',
                'title': 'Ask a question',
                'message': 'Type a question or choose an example prompt.',
            }
        elif "summary" in lower or "how many" in lower or "total" in lower:
            st.session_state.dual_ai_answer = {
                'kind': 'summary',
                'title': 'Allocation Summary',
                'summary': {
                    'Total Students': st.session_state.stats.total_students,
                    'Allocated': st.session_state.stats.allocated_students,
                    'Unallocated': st.session_state.stats.unallocated_students,
                    'Success Rate %': round(st.session_state.stats.allocation_rate * 100, 2),
                },
            }
        elif "unallocated" in lower or "waitlist" in lower or "not allocated" in lower:
            st.session_state.dual_ai_answer = {
                'kind': 'table',
                'title': 'Unallocated Students',
                'message': f'Found {len(unallocated)} unallocated students.',
                'rows': [
                    {
                        'Name': r.name,
                        'Registration': r.registration_number,
                        'Reason': r.reason,
                    }
                    for r in unallocated
                ],
            }
        else:
            section_match = re.search(r"\b(?:section|sec)\s*(\d+)\b", lower)
            section_num = section_match.group(1) if section_match else None
            course_names = sorted({
                *(r.g1_course for r in allocated if r.g1_course),
                *(r.g2_course for r in allocated if r.g2_course),
            })

            matched_course = None
            for name in course_names:
                if str(name).lower() in lower:
                    matched_course = name
                    break

            filtered = allocated
            if section_num:
                filtered = [r for r in filtered if f"section {section_num}" in str(r.pair_label or "").lower()]
            if matched_course:
                filtered = [
                    r for r in filtered
                    if str(r.g1_course or "").lower() == matched_course.lower()
                    or str(r.g2_course or "").lower() == matched_course.lower()
                ]

            if filtered:
                st.session_state.dual_ai_answer = {
                    'kind': 'table',
                    'title': 'Filtered Allocated Students',
                    'message': f'Matched {len(filtered)} students.',
                    'rows': [
                        {
                            'Name': r.name,
                            'Registration': r.registration_number,
                            'G1 Course': r.g1_course,
                            'G2 Course': r.g2_course,
                            'Section': r.pair_label,
                            'G1 Rank': (r.g1_preference_rank + 1) if r.g1_preference_rank is not None else 'N/A',
                            'G2 Rank': (r.g2_preference_rank + 1) if r.g2_preference_rank is not None else 'N/A',
                        }
                        for r in filtered
                    ],
                }
            else:
                st.session_state.dual_ai_answer = {
                    'kind': 'text',
                    'title': 'No direct match found',
                    'message': 'Try prompts like: Show allocated students in Section 1, List unallocated students, or Give me allocation summary.',
                }

    dual_ai_answer = st.session_state.get('dual_ai_answer')
    if dual_ai_answer:
        st.markdown(f"**{dual_ai_answer.get('title', 'AI Response')}**")
        if dual_ai_answer.get('message'):
            st.write(dual_ai_answer['message'])

        if dual_ai_answer.get('kind') == 'summary':
            summary = dual_ai_answer.get('summary', {})
            cols = st.columns(max(len(summary), 1))
            for idx, (label, value) in enumerate(summary.items()):
                with cols[idx]:
                    st.metric(label, value)
        elif dual_ai_answer.get('kind') == 'table':
            st.dataframe(pd.DataFrame(dual_ai_answer.get('rows', [])), use_container_width=True)
    
    st.divider()
    
    # Allocated Students
    st.subheader("✅ Allocated Students")
    if allocated:
        df_allocated = pd.DataFrame([{
            'Name': r.name,
            'Registration': r.registration_number,
            'G1 Course': r.g1_course,
            'G2 Course': r.g2_course,
            'Section': r.pair_label,
            'G1 Rank': r.g1_preference_rank + 1,
            'G2 Rank': r.g2_preference_rank + 1,
            'Fallback': '✓' if r.fallback_used else '✗'
        } for r in allocated])
        
        st.dataframe(df_allocated, use_container_width=True)
        
        # Download allocated
        csv_allocated = df_allocated.to_csv(index=False)
        st.download_button(
            label="📥 Download Allocated Students (CSV)",
            data=csv_allocated,
            file_name=f"allocated_students_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
    else:
        st.info("No students allocated.")
    
    st.divider()
    
    # Unallocated Students
    st.subheader("❌ Unallocated Students")
    if unallocated:
        df_unallocated = pd.DataFrame([{
            'Name': r.name,
            'Registration': r.registration_number,
            'Reason': r.reason
        } for r in unallocated])
        
        st.dataframe(df_unallocated, use_container_width=True)
        
        # Download unallocated
        csv_unallocated = df_unallocated.to_csv(index=False)
        st.download_button(
            label="📥 Download Unallocated Students (CSV)",
            data=csv_unallocated,
            file_name=f"unallocated_students_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
    else:
        st.success("✅ All students successfully allocated!")
    
    st.divider()
    
    # Section-wise breakdown
    st.subheader("📊 Section-wise Breakdown")
    section_groups = {}
    for result in allocated:
        section = result.pair_label
        if section not in section_groups:
            section_groups[section] = []
        section_groups[section].append(result)
    
    for section, students in sorted(section_groups.items()):
        with st.expander(f"{section} ({len(students)} students)"):
            df_section = pd.DataFrame([{
                'Name': s.name,
                'Registration': s.registration_number,
                'G1 Rank': s.g1_preference_rank + 1,
                'G2 Rank': s.g2_preference_rank + 1
            } for s in students])
            st.dataframe(df_section, use_container_width=True)
    
    # ==== COMPREHENSIVE DOWNLOAD REPORT ====
    st.divider()
    st.subheader("📋 Export Complete Report")
    st.info("Download comprehensive Excel report with all allocation details")
    
    # Generate Excel export
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        excel_buffer = BytesIO()
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # ===== SHEET 1: STATISTICS =====
        ws_stats = wb.create_sheet("Statistics")
        ws_stats['A1'] = "Allocation Statistics"
        ws_stats['A1'].font = Font(bold=True, size=12)
        
        row = 3
        ws_stats[f'A{row}'] = "Total Students"
        ws_stats[f'B{row}'] = len(st.session_state.results)
        row += 1
        
        ws_stats[f'A{row}'] = "Allocated Students"
        ws_stats[f'B{row}'] = len(allocated)
        row += 1
        
        ws_stats[f'A{row}'] = "Unallocated Students"
        ws_stats[f'B{row}'] = len(unallocated)
        row += 1
        
        if len(st.session_state.results) > 0:
            allocation_rate = (len(allocated) / len(st.session_state.results)) * 100
            ws_stats[f'A{row}'] = "Allocation Rate (%)"
            ws_stats[f'B{row}'] = f"{allocation_rate:.2f}%"
            row += 1
        
        if st.session_state.stats:
            ws_stats[f'A{row}'] = "Allocation Criteria"
            ws_stats[f'B{row}'] = st.session_state.allocation_criteria
            row += 1
            
            ws_stats[f'A{row}'] = "Section Mode"
            mode = "Automatic (Edited)" if st.session_state.auto_sections_edited else st.session_state.section_mode
            ws_stats[f'B{row}'] = mode
            row += 1
            
            if st.session_state.section_mode == "automatic":
                ws_stats[f'A{row}'] = "Sections Created"
                ws_stats[f'B{row}'] = st.session_state.auto_num_sections
                row += 1
        
        # ===== SHEET 2: ALLOCATED STUDENTS =====
        ws_allocated = wb.create_sheet("Allocated Students")
        headers_allocated = ['Name', 'Registration', 'G1 Course', 'G2 Course', 'Section', 'G1 Rank', 'G2 Rank', 'Fallback']
        for col_idx, header in enumerate(headers_allocated, 1):
            cell = ws_allocated.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for row_idx, r in enumerate(allocated, 2):
            ws_allocated.cell(row=row_idx, column=1).value = r.name
            ws_allocated.cell(row=row_idx, column=2).value = r.registration_number
            ws_allocated.cell(row=row_idx, column=3).value = r.g1_course
            ws_allocated.cell(row=row_idx, column=4).value = r.g2_course
            ws_allocated.cell(row=row_idx, column=5).value = r.pair_label
            ws_allocated.cell(row=row_idx, column=6).value = r.g1_preference_rank + 1
            ws_allocated.cell(row=row_idx, column=7).value = r.g2_preference_rank + 1
            ws_allocated.cell(row=row_idx, column=8).value = 'Yes' if r.fallback_used else 'No'
        
        # Auto-width columns
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws_allocated.column_dimensions[col].width = 18
        
        # ===== SHEET 3: UNALLOCATED STUDENTS =====
        ws_unallocated = wb.create_sheet("Unallocated Students")
        headers_unallocated = ['Name', 'Registration', 'Reason']
        for col_idx, header in enumerate(headers_unallocated, 1):
            cell = ws_unallocated.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        
        for row_idx, r in enumerate(unallocated, 2):
            ws_unallocated.cell(row=row_idx, column=1).value = r.name
            ws_unallocated.cell(row=row_idx, column=2).value = r.registration_number
            ws_unallocated.cell(row=row_idx, column=3).value = r.reason
        
        # Auto-width columns
        for col in ['A', 'B', 'C']:
            ws_unallocated.column_dimensions[col].width = 25
        
        # ===== SHEET 4+: SECTION-WISE BREAKDOWN =====
        for section_num, (section, students) in enumerate(sorted(section_groups.items()), 1):
            ws_section = wb.create_sheet(f"Section {section_num}")
            
            # Section header
            ws_section['A1'] = f"Section: {section}"
            ws_section['A1'].font = Font(bold=True, size=12)
            ws_section['A2'] = f"Total Students: {len(students)}"
            
            # Headers
            headers_section = ['Name', 'Registration', 'G1 Rank', 'G2 Rank']
            for col_idx, header in enumerate(headers_section, 1):
                cell = ws_section.cell(row=4, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            
            # Data
            for row_idx, student in enumerate(students, 5):
                ws_section.cell(row=row_idx, column=1).value = student.name
                ws_section.cell(row=row_idx, column=2).value = student.registration_number
                ws_section.cell(row=row_idx, column=3).value = student.g1_preference_rank + 1
                ws_section.cell(row=row_idx, column=4).value = student.g2_preference_rank + 1
            
            # Auto-width columns
            for col in ['A', 'B', 'C', 'D']:
                ws_section.column_dimensions[col].width = 18
        
        # Save to buffer
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Download button
        st.download_button(
            label="📊 Download Complete Report (Excel)",
            data=excel_buffer.getvalue(),
            file_name=f"allocation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        st.error(f"Error generating report: {str(e)}")
    
    # Action buttons
    st.divider()
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("🔄 Back to Review Sections"):
            st.session_state.step = 2
            st.rerun()
    
    with col2:
        if st.button("🏠 Start Over"):
            st.session_state.step = 1
            st.session_state.results = None
            st.session_state.stats = None
            st.session_state.paired_sections = []
            st.rerun()
