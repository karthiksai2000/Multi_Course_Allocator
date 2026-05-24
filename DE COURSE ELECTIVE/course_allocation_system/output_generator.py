import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List
import os

class OutputGenerator:
    """Generate Excel output with results"""
    
    @staticmethod
    def create_output_file(allocation_result: Dict, output_path: str):
        """Create comprehensive Excel output"""
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create sheets
        OutputGenerator._create_allocated_sheet(wb, allocation_result)
        OutputGenerator._create_wait_listed_special_sheet(wb, allocation_result)
        OutputGenerator._create_waitlisted_regular_sheet(wb, allocation_result)
        OutputGenerator._create_waitlist_sheet(wb, allocation_result)
        OutputGenerator._create_statistics_sheet(wb, allocation_result)
        
        wb.save(output_path)
        return output_path
    
    @staticmethod
    def _style_header(ws, row, col_count):
        """Apply header styling"""
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
    
    @staticmethod
    def _apply_border(ws, start_row, end_row, start_col, end_col):
        """Apply borders to range"""
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = border
    
    @staticmethod
    def _create_allocated_sheet(wb, allocation_result: Dict):
        """Create allocated students sheet"""
        ws = wb.create_sheet('Allocated Students')
        
        # Headers
        headers = ['S.No', 'Name', 'Reg No', 'CGPA', 'Allocated Course', 'Section', 'Preference Rank', 'Reason', 'Student Type']
        
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        OutputGenerator._style_header(ws, 1, len(headers))
        
        # Data - Separate regular and research students
        row = 2
        allocated_data = allocation_result.get('allocated_students', [])
        
        # Add regular students first
        regular = [s for s in allocated_data if s.get('student_type') == 'Regular']
        for idx, student in enumerate(regular, 1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=student.get('name', 'N/A'))
            ws.cell(row=row, column=3, value=student.get('reg_no', 'N/A'))
            ws.cell(row=row, column=4, value=student.get('cgpa', 0))
            ws.cell(row=row, column=5, value=student.get('allocated_course', 'N/A'))
            
            # Find section
            section_id = 'N/A'
            course_name = student.get('allocated_course')
            if course_name and 'allocation_section_id' in student:
                section_id = f"Section {student['allocation_section_id']}"
            
            ws.cell(row=row, column=6, value=section_id)
            ws.cell(row=row, column=7, value=student.get('preference_rank', 'N/A'))
            ws.cell(row=row, column=8, value=student.get('allocation_reason', 'Successfully allocated'))
            ws.cell(row=row, column=9, value='Regular')
            row += 1
        
        regular_count = len(regular)
        
        # Add research students
        research = [s for s in allocated_data if s.get('student_type') == 'Research/Special']
        for idx, student in enumerate(research, 1):
            ws.cell(row=row, column=1, value=f"R{idx}")
            ws.cell(row=row, column=2, value=student.get('name', 'N/A'))
            ws.cell(row=row, column=3, value=student.get('reg_no', 'N/A'))
            ws.cell(row=row, column=4, value=student.get('cgpa', 0))
            ws.cell(row=row, column=5, value=student.get('allocated_course', 'N/A'))
            
            section_id = 'N/A'
            if 'allocation_section_id' in student:
                section_id = f"Section {student['allocation_section_id']}"
            
            ws.cell(row=row, column=6, value=section_id)
            ws.cell(row=row, column=7, value=student.get('preference_rank', 'Special'))
            ws.cell(row=row, column=8, value=student.get('allocation_reason', 'Allocated to special section'))
            ws.cell(row=row, column=9, value='Research/Special')
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 40  # Wider for reason
        ws.column_dimensions['I'].width = 20
        
        # Set text wrapping for reason column
        for row_idx in range(2, row):
            ws.cell(row=row_idx, column=8).alignment = Alignment(wrap_text=True)
        
        OutputGenerator._apply_border(ws, 1, row - 1, 1, len(headers))
        
        # Add summary
        summary_row = row + 2
        ws.cell(row=summary_row, column=1, value='Allocated Summary:')
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        
        ws.cell(row=summary_row + 1, column=1, value='Regular Students Allocated:')
        ws.cell(row=summary_row + 1, column=2, value=regular_count)
        
        ws.cell(row=summary_row + 2, column=1, value='Research Students Allocated:')
        ws.cell(row=summary_row + 2, column=2, value=len(research))
        
        ws.cell(row=summary_row + 3, column=1, value='Total Allocated:')
        ws.cell(row=summary_row + 3, column=2, value=len(allocated_data))
        ws.cell(row=summary_row + 3, column=2).font = Font(bold=True)
    
    @staticmethod
    def _create_wait_listed_special_sheet(wb, allocation_result: Dict):
        """Create waitlist sheet for SPECIAL/RESEARCH students only"""
        ws = wb.create_sheet('Waitlist - Special Students', 1)
        
        headers = ['S.No', 'Name', 'Reg No', 'CGPA', 'Reason for Waitlist']
        
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        OutputGenerator._style_header(ws, 1, len(headers))
        
        waitlist = allocation_result.get('waitlist', [])
        special_waitlist = [s for s in waitlist if s.get('student_type') == 'Research/Special']
        
        for idx, student in enumerate(special_waitlist, 1):
            ws.cell(row=idx + 1, column=1, value=idx)
            ws.cell(row=idx + 1, column=2, value=student.get('name', 'N/A'))
            ws.cell(row=idx + 1, column=3, value=student.get('reg_no', 'N/A'))
            ws.cell(row=idx + 1, column=4, value=student.get('cgpa', 0))
            
            # Add reason for waitlisting
            reason = student.get('allocation_reason', 'Unknown reason')
            ws.cell(row=idx + 1, column=5, value=reason)
            
            # Set text wrapping for reason column
            ws.cell(row=idx + 1, column=5).alignment = Alignment(wrap_text=True)
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 60  # Wider for reason
        
        OutputGenerator._apply_border(ws, 1, len(special_waitlist) + 1, 1, len(headers))
        
        # Summary
        summary_row = len(special_waitlist) + 3
        ws.cell(row=summary_row, column=1, value='Total Waitlisted (Special):')
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=len(special_waitlist))
        ws.cell(row=summary_row, column=2).font = Font(bold=True)
    
    @staticmethod
    def _create_waitlisted_regular_sheet(wb, allocation_result: Dict):
        """Create waitlist sheet for REGULAR students only"""
        ws = wb.create_sheet('Waitlist - Regular Students', 2)
        
        headers = ['S.No', 'Name', 'Reg No', 'CGPA', 'Preferences', 'Reason for Waitlist']
        
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        OutputGenerator._style_header(ws, 1, len(headers))
        
        waitlist = allocation_result.get('waitlist', [])
        regular_waitlist = [s for s in waitlist if s.get('student_type') == 'Regular']
        
        for idx, student in enumerate(regular_waitlist, 1):
            ws.cell(row=idx + 1, column=1, value=idx)
            ws.cell(row=idx + 1, column=2, value=student.get('name', 'N/A'))
            ws.cell(row=idx + 1, column=3, value=student.get('reg_no', 'N/A'))
            ws.cell(row=idx + 1, column=4, value=student.get('cgpa', 0))
            
            prefs = ', '.join(student.get('preferences', []))
            ws.cell(row=idx + 1, column=5, value=prefs)
            
            # Add reason for waitlisting
            reason = student.get('allocation_reason', 'Unknown reason')
            ws.cell(row=idx + 1, column=6, value=reason)
            
            # Set text wrapping for reason column
            ws.cell(row=idx + 1, column=6).alignment = Alignment(wrap_text=True)
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 60  # Wider for reason
        
        OutputGenerator._apply_border(ws, 1, len(regular_waitlist) + 1, 1, len(headers))
        
        # Summary
        summary_row = len(regular_waitlist) + 3
        ws.cell(row=summary_row, column=1, value='Total Waitlisted (Regular):')
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=len(regular_waitlist))
        ws.cell(row=summary_row, column=2).font = Font(bold=True)
    
    @staticmethod
    def _create_waitlist_sheet(wb, allocation_result: Dict):
        """Create combined waitlist sheet with all unallocated students"""
        ws = wb.create_sheet('Waitlist - All', 3)
        
        headers = ['S.No', 'Name', 'Reg No', 'CGPA', 'Type', 'Preferences', 'Reason for Waitlist']
        
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        OutputGenerator._style_header(ws, 1, len(headers))
        
        waitlist = allocation_result.get('waitlist', [])
        
        # Separate special and regular for grouping
        special_waitlist = [s for s in waitlist if s.get('student_type') == 'Research/Special']
        regular_waitlist = [s for s in waitlist if s.get('student_type') == 'Regular']
        
        row = 2
        idx = 1
        
        # First, add all special students
        if special_waitlist:
            for student in special_waitlist:
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=student.get('name', 'N/A'))
                ws.cell(row=row, column=3, value=student.get('reg_no', 'N/A'))
                ws.cell(row=row, column=4, value=student.get('cgpa', 0))
                ws.cell(row=row, column=5, value='🔬 SPECIAL')
                ws.cell(row=row, column=6, value='N/A')  # Special students have no preferences
                
                reason = student.get('allocation_reason', 'Unknown reason')
                ws.cell(row=row, column=7, value=reason)
                ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True)
                
                # Color special students differently
                fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = fill
                
                row += 1
                idx += 1
        
        # Then add all regular students
        if regular_waitlist:
            for student in regular_waitlist:
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=student.get('name', 'N/A'))
                ws.cell(row=row, column=3, value=student.get('reg_no', 'N/A'))
                ws.cell(row=row, column=4, value=student.get('cgpa', 0))
                ws.cell(row=row, column=5, value='📚 REGULAR')
                
                prefs = ', '.join(student.get('preferences', []))
                ws.cell(row=row, column=6, value=prefs)
                
                reason = student.get('allocation_reason', 'Unknown reason')
                ws.cell(row=row, column=7, value=reason)
                ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True)
                
                row += 1
                idx += 1
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 70  # Wider for reason
        
        OutputGenerator._apply_border(ws, 1, len(waitlist) + 1, 1, len(headers))
        
        # Summary
        summary_row = len(waitlist) + 3
        ws.cell(row=summary_row, column=1, value='Total Waitlisted:')
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=len(waitlist))
        ws.cell(row=summary_row, column=2).font = Font(bold=True)
        
        ws.cell(row=summary_row + 1, column=1, value='Special/Research:')
        ws.cell(row=summary_row + 1, column=2, value=len(special_waitlist))
        
        ws.cell(row=summary_row + 2, column=1, value='Regular:')
        ws.cell(row=summary_row + 2, column=2, value=len(regular_waitlist))
    
    @staticmethod
    def _create_statistics_sheet(wb, allocation_result: Dict):
        """Create statistics sheet"""
        ws = wb.create_sheet('Statistics')
        
        stats = allocation_result.get('statistics', {})
        
        # Overall statistics
        row = 1
        ws.cell(row=row, column=1, value='Overall Statistics').font = Font(bold=True, size=12)
        
        row = 2
        ws.cell(row=row, column=1, value='Total Students:')
        ws.cell(row=row, column=2, value=stats.get('total_students', 0))
        
        row += 1
        ws.cell(row=row, column=1, value='Regular Students:')
        ws.cell(row=row, column=2, value=stats.get('regular_students', 0))
        
        row += 1
        ws.cell(row=row, column=1, value='Research Students:')
        ws.cell(row=row, column=2, value=stats.get('research_students', 0))
        
        row += 1
        ws.cell(row=row, column=1, value='Allocated:')
        ws.cell(row=row, column=2, value=stats.get('allocated', 0))
        ws.cell(row=row, column=2).font = Font(bold=True, color='008000')
        
        row += 1
        ws.cell(row=row, column=1, value='Waitlisted:')
        ws.cell(row=row, column=2, value=stats.get('waitlisted', 0))
        ws.cell(row=row, column=2).font = Font(bold=True, color='FF0000')
        
        # Allocation success rate
        row += 2
        total = stats.get('total_students', 1)
        allocated = stats.get('allocated', 0)
        success_rate = (allocated / total * 100) if total > 0 else 0
        
        ws.cell(row=row, column=1, value='Allocation Success Rate:')
        ws.cell(row=row, column=2, value=f"{success_rate:.2f}%")
        ws.cell(row=row, column=2).font = Font(bold=True, size=11)
        
        # Course-wise statistics
        row += 3
        ws.cell(row=row, column=1, value='Course-wise Statistics').font = Font(bold=True, size=12)
        
        headers = ['Course', 'Sections', 'Capacity', 'Allocated', 'Unfilled', 'Utilization %']
        row += 1
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row, column=col_idx, value=header)
        OutputGenerator._style_header(ws, row, len(headers))
        
        courses = stats.get('courses', {})
        row += 1
        start_row = row
        
        for course_name, course_stats in courses.items():
            ws.cell(row=row, column=1, value=course_name)
            ws.cell(row=row, column=2, value=course_stats.get('sections', 0))
            ws.cell(row=row, column=3, value=course_stats.get('total_capacity', 0))
            ws.cell(row=row, column=4, value=course_stats.get('total_allocated', 0))
            ws.cell(row=row, column=5, value=course_stats.get('unfilled', 0))
            ws.cell(row=row, column=6, value=f"{course_stats.get('utilization_percent', 0):.2f}%")
            row += 1
        
        OutputGenerator._apply_border(ws, start_row - 1, row - 1, 1, len(headers))
        
        # Section details
        if courses:
            row += 2
            ws.cell(row=row, column=1, value='Section Details').font = Font(bold=True, size=12)
            
            for course_name, course_stats in courses.items():
                row += 1
                ws.cell(row=row, column=1, value=course_name).font = Font(bold=True)
                
                headers = ['Section', 'Capacity', 'Allocated', 'Available']
                row += 1
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col_idx, value=header)
                OutputGenerator._style_header(ws, row, len(headers))
                
                section_start = row + 1
                for section in course_stats.get('section_details', []):
                    row += 1
                    ws.cell(row=row, column=1, value=section['section'])
                    ws.cell(row=row, column=2, value=section['capacity'])
                    ws.cell(row=row, column=3, value=section['allocated'])
                    ws.cell(row=row, column=4, value=section['available'])
                
                OutputGenerator._apply_border(ws, section_start - 1, row, 1, len(headers))
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
